"""
Spreadsheet exporter module for test framework.

This module provides advanced functionality for exporting test
results to various spreadsheet formats with comprehensive metrics.
"""

import os
import json
import logging
import re
from typing import Dict, List, Any, Optional, Set, Tuple
from dataclasses import dataclass
from datetime import datetime
from collections import defaultdict

from rv_android_core.util.logging.constants import CONTEXT_COMPONENT
from rv_android_core.util.logging.manager import LoggingManager


class SpreadsheetExporter:
    """
    Advanced exporter for test results to spreadsheet formats.
    
    Generates detailed spreadsheet exports with comprehensive metrics,
    supporting both CSV and Excel formats with multiple sheets
    for different analysis perspectives.
    
    ### Key Responsibilities:
    - Exports comprehensive test results to CSV and Excel formats
    - Creates multiple sheets for different analysis perspectives
    - Supports detailed metrics and statistical analysis
    - Enables trend analysis and comparison between configurations
    """
    
    def __init__(self):
        """Initialize the spreadsheet exporter."""
        # Set up logging
        self.logger = LoggingManager.get_instance().get_logger(
            'test_framework.spreadsheet_exporter',
            {CONTEXT_COMPONENT: 'SpreadsheetExporter'}
        )
        
        # Define common metrics to export
        self.common_metrics = [
            'avg_method_coverage',
            'avg_activity_coverage', 
            'avg_mop_method_coverage',
            'avg_execution_time',
            'overall_score'
        ]
        
        # Define metric display names
        self.metric_display_names = {
            'avg_method_coverage': 'Method Coverage (%)',
            'avg_activity_coverage': 'Activity Coverage (%)',
            'avg_mop_method_coverage': 'MOP Method Coverage (%)',
            'avg_execution_time': 'Execution Time (s)',
            'overall_score': 'Overall Score'
        }
    
    def export_to_csv(self, results: Dict[str, Any], output_file: str) -> bool:
        """
        Export analysis results to CSV format.
        
        Args:
            results: Analysis results dictionary
            output_file: Path to save the CSV file
            
        Returns:
            True if export succeeded, False otherwise
        """
        try:
            import csv
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            # Extract comparisons
            comparisons = results.get('configuration_comparisons', {})
            
            # No data to export
            if not comparisons:
                self.logger.warning("No comparison data to export")
                return False
            
            # Prepare configuration data
            config_data = self._prepare_config_data(comparisons)
            
            # Export by configuration
            self._export_configs_to_csv(config_data, output_file)
            
            # Export by app if app data is available
            app_data = self._prepare_app_data(results)
            if app_data:
                app_output_file = os.path.splitext(output_file)[0] + "_by_app.csv"
                self._export_apps_to_csv(app_data, app_output_file)
            
            # Export tool comparison if multiple tools are present
            tool_data = self._prepare_tool_data(comparisons)
            if len(tool_data) > 1:
                tool_output_file = os.path.splitext(output_file)[0] + "_by_tool.csv"
                self._export_tools_to_csv(tool_data, tool_output_file)
            
            # Export correlation data if available
            if results.get('correlation_report'):
                correlation_output_file = os.path.splitext(output_file)[0] + "_correlations.csv"
                self._export_correlations_to_csv(results['correlation_report'], correlation_output_file)
            
            # Export anomaly data if available
            if results.get('anomaly_report'):
                anomaly_output_file = os.path.splitext(output_file)[0] + "_anomalies.csv"
                self._export_anomalies_to_csv(results['anomaly_report'], anomaly_output_file)
            
            self.logger.info(f"Successfully exported results to CSV: {output_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error exporting to CSV: {str(e)}")
            return False
    
    def export_to_excel(self, results: Dict[str, Any], output_file: str) -> bool:
        """
        Export analysis results to Excel format with multiple sheets.
        
        Args:
            results: Analysis results dictionary
            output_file: Path to save the Excel file
            
        Returns:
            True if export succeeded, False otherwise
        """
        try:
            # Import pandas and openpyxl
            import pandas as pd
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            # Extract comparisons
            comparisons = results.get('configuration_comparisons', {})
            
            # No data to export
            if not comparisons:
                self.logger.warning("No comparison data to export")
                return False
            
            # Prepare data
            config_data = self._prepare_config_data(comparisons)
            app_data = self._prepare_app_data(results)
            tool_data = self._prepare_tool_data(comparisons)
            
            # Create Excel writer
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Create configuration sheet
                configs_df = pd.DataFrame(config_data)
                configs_df.to_excel(writer, sheet_name='Configurations', index=False)
                self._format_excel_sheet(writer, 'Configurations', configs_df)
                
                # Create summary sheet
                summary_data = {
                    'Metric': ['Total Configurations', 'Total Applications', 'Analysis Date'],
                    'Value': [
                        len(config_data),
                        results.get('total_apps', 0),
                        datetime.now().strftime('%Y-%m-%d %H:%M')
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                self._format_excel_sheet(writer, 'Summary', summary_df)
                
                # Create top configurations sheet
                top_configs = results.get('top_configurations', {})
                top_data = []
                
                for metric, config_ids in top_configs.items():
                    for rank, config_id in enumerate(config_ids[:5], 1):
                        config = next((c for c in config_data if c['Configuration ID'] == config_id), {})
                        if config:
                            metric_display = self.metric_display_names.get(metric, metric)
                            metric_value = config.get(metric_display, 0)
                            
                            top_data.append({
                                'Metric': metric_display,
                                'Rank': rank,
                                'Configuration ID': config_id,
                                'Value': metric_value
                            })
                
                if top_data:
                    top_df = pd.DataFrame(top_data)
                    top_df.to_excel(writer, sheet_name='Top Configurations', index=False)
                    self._format_excel_sheet(writer, 'Top Configurations', top_df)
                
                # Create app sheet if app data is available
                if app_data:
                    apps_df = pd.DataFrame(app_data)
                    apps_df.to_excel(writer, sheet_name='By App', index=False)
                    self._format_excel_sheet(writer, 'By App', apps_df)
                
                # Create tool comparison sheet if multiple tools
                if len(tool_data) > 1:
                    tools_df = pd.DataFrame(tool_data)
                    tools_df.to_excel(writer, sheet_name='By Tool', index=False)
                    self._format_excel_sheet(writer, 'By Tool', tools_df)
                
                # Create correlation sheet if available
                if results.get('correlation_report'):
                    corr_data = self._prepare_correlation_data(results['correlation_report'])
                    if corr_data:
                        corr_df = pd.DataFrame(corr_data)
                        corr_df.to_excel(writer, sheet_name='Correlations', index=False)
                        self._format_excel_sheet(writer, 'Correlations', corr_df)
                
                # Create anomalies sheet if available
                if results.get('anomaly_report'):
                    anomaly_data = self._prepare_anomaly_data(results['anomaly_report'])
                    if anomaly_data:
                        anomaly_df = pd.DataFrame(anomaly_data)
                        anomaly_df.to_excel(writer, sheet_name='Anomalies', index=False)
                        self._format_excel_sheet(writer, 'Anomalies', anomaly_df)
                
                # Create recommendations sheet if available
                if results.get('correlation_report') and results['correlation_report'].get('recommendations'):
                    rec_data = self._prepare_recommendation_data(results['correlation_report'])
                    if rec_data:
                        rec_df = pd.DataFrame(rec_data)
                        rec_df.to_excel(writer, sheet_name='Recommendations', index=False)
                        self._format_excel_sheet(writer, 'Recommendations', rec_df)
            
            self.logger.info(f"Successfully exported results to Excel: {output_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {str(e)}")
            return False
    
    def _prepare_config_data(self, comparisons: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Prepare configuration data for export.
        
        Args:
            comparisons: Configuration comparison data
            
        Returns:
            List of dictionaries with configuration data
        """
        config_data = []
        
        for config_id, data in comparisons.items():
            row = {'Configuration ID': config_id}
            
            # Extract tool and other components from config_id
            if '_' in config_id:
                parts = config_id.split('_')
                row['Tool'] = parts[0]
                if len(parts) >= 2:
                    row['LLM Type'] = parts[1]
                    if len(parts) >= 3:
                        row['LLM Model'] = '_'.join(parts[2:])
            
            # Add metrics
            avg_metrics = data.get('avg_metrics', {})
            for metric in self.common_metrics:
                if metric in avg_metrics:
                    row[self.metric_display_names.get(metric, metric)] = avg_metrics[metric]
            
            # Add counts
            row['App Count'] = data.get('app_count', 0)
            row['Error Count'] = data.get('error_count', 0)
            row['Result Count'] = data.get('result_count', 0)
            
            # Add success rate
            if data.get('app_count', 0) > 0:
                error_count = data.get('error_count', 0)
                app_count = data.get('app_count', 0)
                success_rate = 100.0 * (app_count - error_count) / app_count
                row['Success Rate (%)'] = success_rate
            
            config_data.append(row)
        
        return config_data
    
    def _prepare_app_data(self, results: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Prepare app-specific data for export.
        
        Args:
            results: Analysis results dictionary
            
        Returns:
            List of dictionaries with app data
        """
        app_data = []
        
        # Check if app metrics are available
        app_metrics = results.get('app_metrics', {})
        if not app_metrics:
            return app_data
        
        # Prepare rows for each app
        for app_name, app_info in app_metrics.items():
            row = {'App Name': app_name}
            
            # Add basic metrics
            for metric in ['method_coverage', 'activity_coverage', 'mop_method_coverage']:
                if metric in app_info:
                    display_name = metric.replace('_', ' ').title()
                    row[display_name] = app_info[metric]
            
            # Add counts if available
            for count_metric in ['activity_count', 'method_count', 'mop_spec_count']:
                if count_metric in app_info:
                    display_name = count_metric.replace('_', ' ').title()
                    row[display_name] = app_info[count_metric]
            
            # Add top performing configurations for this app
            config_metrics = app_info.get('config_metrics', {})
            if config_metrics:
                # Find best config for overall score
                best_configs = []
                for metric in self.common_metrics:
                    metric_values = {
                        config_id: metrics.get(metric, 0) 
                        for config_id, metrics in config_metrics.items()
                    }
                    
                    if metric_values:
                        # Sort by value (descending for all except execution time)
                        reverse = (metric != 'avg_execution_time')
                        sorted_configs = sorted(
                            metric_values.items(), 
                            key=lambda x: x[1], 
                            reverse=reverse
                        )
                        
                        # Take top config
                        if sorted_configs:
                            best_config, best_value = sorted_configs[0]
                            display_name = f"Best {self.metric_display_names.get(metric, metric)}"
                            row[display_name] = best_config
                            row[f"{display_name} Value"] = best_value
                            best_configs.append(best_config)
                
                # Count unique best configs
                row['Unique Best Configs'] = len(set(best_configs))
            
            app_data.append(row)
        
        return app_data
    
    def _prepare_tool_data(self, comparisons: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Prepare tool-specific data for export.
        
        Args:
            comparisons: Configuration comparison data
            
        Returns:
            List of dictionaries with tool data
        """
        tool_metrics = defaultdict(lambda: {
            'configs': 0,
            'apps': 0,
            'errors': 0,
            'metrics': defaultdict(list)
        })
        
        for config_id, data in comparisons.items():
            # Extract tool from config_id
            tool = config_id.split('_')[0] if '_' in config_id else 'unknown'
            
            # Update counts
            tool_metrics[tool]['configs'] += 1
            tool_metrics[tool]['apps'] += data.get('app_count', 0)
            tool_metrics[tool]['errors'] += data.get('error_count', 0)
            
            # Update metrics
            avg_metrics = data.get('avg_metrics', {})
            for metric in self.common_metrics:
                if metric in avg_metrics:
                    tool_metrics[tool]['metrics'][metric].append(avg_metrics[metric])
        
        # Calculate averages
        tool_data = []
        for tool, data in tool_metrics.items():
            row = {'Tool': tool}
            row['Configuration Count'] = data['configs']
            row['Total Apps Tested'] = data['apps']
            row['Total Errors'] = data['errors']
            
            # Calculate success rate
            if data['apps'] > 0:
                success_rate = 100.0 * (data['apps'] - data['errors']) / data['apps']
                row['Success Rate (%)'] = success_rate
            
            # Calculate metric averages
            for metric, values in data['metrics'].items():
                if values:
                    avg_value = sum(values) / len(values)
                    row[self.metric_display_names.get(metric, metric)] = avg_value
            
            tool_data.append(row)
        
        return tool_data
    
    def _prepare_correlation_data(self, correlation_report: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Prepare correlation data for export.
        
        Args:
            correlation_report: Correlation analysis report
            
        Returns:
            List of dictionaries with correlation data
        """
        correlation_data = []
        
        # Get top correlations
        top_correlations = correlation_report.get('top_correlations', [])
        
        for corr in top_correlations:
            row = {
                'App Characteristic': corr.get('app_characteristic', '').replace('_', ' ').title(),
                'Configuration ID': corr.get('config_id', ''),
                'Metric': corr.get('config_metric', '').replace('avg_', '').replace('_', ' ').title(),
                'Correlation Value': corr.get('correlation_value', 0),
                'Sample Size': corr.get('sample_size', 0),
                'Confidence': corr.get('confidence', '').title(),
            }
            
            if corr.get('p_value') is not None:
                row['P-Value'] = corr['p_value']
                
            correlation_data.append(row)
        
        return correlation_data
    
    def _prepare_anomaly_data(self, anomaly_report: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Prepare anomaly data for export.
        
        Args:
            anomaly_report: Anomaly detection report
            
        Returns:
            List of dictionaries with anomaly data
        """
        anomaly_data = []
        
        # Get anomalies
        anomalies = anomaly_report.get('anomalies', [])
        
        for anomaly in anomalies:
            row = {
                'ID': anomaly.get('id', ''),
                'Type': anomaly.get('type', '').title(),
                'Metric': anomaly.get('metric', '').replace('avg_', '').replace('_', ' ').title(),
                'Expected Value': anomaly.get('expected_value', 0),
                'Actual Value': anomaly.get('actual_value', 0),
                'Deviation (Z-Score)': anomaly.get('deviation', 0),
                'Severity': anomaly.get('severity', '').title(),
            }
                
            anomaly_data.append(row)
        
        return anomaly_data
    
    def _prepare_recommendation_data(self, correlation_report: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Prepare recommendation data for export.
        
        Args:
            correlation_report: Correlation analysis report
            
        Returns:
            List of dictionaries with recommendation data
        """
        recommendation_data = []
        
        # Get recommendations
        recommendations = correlation_report.get('recommendations', {})
        
        for char_name, recs in recommendations.items():
            for rec in recs:
                row = {
                    'App Characteristic': char_name.replace('_', ' ').title(),
                    'Configuration ID': rec.get('config_id', ''),
                    'Correlation': rec.get('correlation', 0),
                    'Confidence': rec.get('confidence', '').title(),
                    'Metric': rec.get('metric', '').replace('avg_', '').replace('_', ' ').title(),
                }
                
                # Add first 100 chars of explanation
                explanation = rec.get('explanation', '')
                if explanation:
                    if len(explanation) > 100:
                        explanation = explanation[:97] + '...'
                    row['Explanation'] = explanation
                    
                recommendation_data.append(row)
        
        return recommendation_data
    
    def _export_configs_to_csv(self, config_data: List[Dict[str, Any]], output_file: str) -> None:
        """
        Export configuration data to CSV.
        
        Args:
            config_data: Prepared configuration data
            output_file: Path to save the CSV file
        """
        import csv
        
        # Determine all fields
        fields = set()
        for row in config_data:
            fields.update(row.keys())
        
        # Ensure important fields come first
        ordered_fields = ['Configuration ID', 'Tool', 'LLM Type', 'LLM Model']
        ordered_fields.extend([f for f in fields if f not in ordered_fields])
        
        # Write CSV
        with open(output_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=ordered_fields)
            writer.writeheader()
            writer.writerows(config_data)
    
    def _export_apps_to_csv(self, app_data: List[Dict[str, Any]], output_file: str) -> None:
        """
        Export app data to CSV.
        
        Args:
            app_data: Prepared app data
            output_file: Path to save the CSV file
        """
        import csv
        
        # Determine all fields
        fields = set()
        for row in app_data:
            fields.update(row.keys())
        
        # Ensure important fields come first
        ordered_fields = ['App Name', 'Method Coverage', 'Activity Coverage', 'Mop Method Coverage']
        ordered_fields.extend([f for f in fields if f not in ordered_fields])
        
        # Write CSV
        with open(output_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=ordered_fields)
            writer.writeheader()
            writer.writerows(app_data)
    
    def _export_tools_to_csv(self, tool_data: List[Dict[str, Any]], output_file: str) -> None:
        """
        Export tool data to CSV.
        
        Args:
            tool_data: Prepared tool data
            output_file: Path to save the CSV file
        """
        import csv
        
        # Determine all fields
        fields = set()
        for row in tool_data:
            fields.update(row.keys())
        
        # Ensure important fields come first
        ordered_fields = ['Tool', 'Configuration Count', 'Total Apps Tested', 'Success Rate (%)']
        ordered_fields.extend([f for f in fields if f not in ordered_fields])
        
        # Write CSV
        with open(output_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=ordered_fields)
            writer.writeheader()
            writer.writerows(tool_data)
    
    def _export_correlations_to_csv(self, correlation_report: Dict[str, Any], output_file: str) -> None:
        """
        Export correlation data to CSV.
        
        Args:
            correlation_report: Correlation analysis report
            output_file: Path to save the CSV file
        """
        import csv
        
        # Prepare correlation data
        correlation_data = self._prepare_correlation_data(correlation_report)
        
        if not correlation_data:
            return
        
        # Determine all fields
        fields = list(correlation_data[0].keys())
        
        # Write CSV
        with open(output_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(correlation_data)
    
    def _export_anomalies_to_csv(self, anomaly_report: Dict[str, Any], output_file: str) -> None:
        """
        Export anomaly data to CSV.
        
        Args:
            anomaly_report: Anomaly detection report
            output_file: Path to save the CSV file
        """
        import csv
        
        # Prepare anomaly data
        anomaly_data = self._prepare_anomaly_data(anomaly_report)
        
        if not anomaly_data:
            return
        
        # Determine all fields
        fields = list(anomaly_data[0].keys())
        
        # Write CSV
        with open(output_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(anomaly_data)
    
    def _format_excel_sheet(self, writer, sheet_name: str, df) -> None:
        """
        Format an Excel sheet for better readability.
        
        Args:
            writer: Excel writer object
            sheet_name: Name of the sheet to format
            df: DataFrame used to populate the sheet
        """
        # Get the worksheet
        worksheet = writer.sheets[sheet_name]
        
        # Set column widths
        for i, col in enumerate(df.columns):
            # Find the longest value in the column
            col_width = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            ) + 2  # Add a little extra space
            
            # Set the column width
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = min(50, col_width)  # Cap at 50 to avoid too wide columns
        
        # Format header
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add borders to header
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.border = thin_border


# Convenient functions
def export_to_enhanced_csv(results: Dict[str, Any], output_file: str) -> bool:
    """
    Export analysis results to enhanced CSV format with multiple files.
    
    Args:
        results: Analysis results dictionary
        output_file: Path to save the main CSV file
        
    Returns:
        True if export succeeded, False otherwise
    """
    exporter = SpreadsheetExporter()
    return exporter.export_to_csv(results, output_file)


def export_to_enhanced_excel(results: Dict[str, Any], output_file: str) -> bool:
    """
    Export analysis results to enhanced Excel format with multiple sheets.
    
    Args:
        results: Analysis results dictionary
        output_file: Path to save the Excel file
        
    Returns:
        True if export succeeded, False otherwise
    """
    exporter = SpreadsheetExporter()
    return exporter.export_to_excel(results, output_file)